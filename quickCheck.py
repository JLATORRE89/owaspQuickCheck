#!/usr/bin/env python3
"""
OWASP QuickCheck - Advanced Security Testing Tool
Complete OWASP Top 10 Coverage with Advanced Features
"""

import requests
import ssl
import socket
import subprocess
import json
import os
import sys
import webbrowser
import argparse
import time
import re
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from urllib.parse import urlparse, urljoin
from typing import Dict, List, Tuple, Any, Optional

# GUI imports
import tkinter as tk
from tkinter import ttk, messagebox, simpledialog

# Advanced imports
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    openpyxl = None

try:
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.lib import colors
except ImportError:
    SimpleDocTemplate = None

try:
    from jinja2 import Template
except ImportError:
    Template = None

try:
    from bs4 import BeautifulSoup
except ImportError:
    BeautifulSoup = None

try:
    from colorama import init, Fore, Style
    init(autoreset=True)
    HAS_COLOR = True
except ImportError:
    HAS_COLOR = False
    class Fore:
        RED = GREEN = YELLOW = CYAN = BLUE = MAGENTA = WHITE = RESET = ""
    class Style:
        BRIGHT = DIM = RESET_ALL = ""

try:
    from tqdm import tqdm
except ImportError:
    tqdm = None

# Import the HackerOne exclusions module
from hackerone_exclusions import should_exclude_test, get_all_exclusions, get_excluded_tests


# ============================================================================
# CONSTANTS AND CONFIGURATIONS
# ============================================================================

USER_AGENT = "OWASP-QuickCheck/2.0 (Security Research)"

# XSS Payloads
XSS_PAYLOADS = [
    "<script>alert('XSS')</script>",
    "<img src=x onerror=alert('XSS')>",
    "<svg/onload=alert('XSS')>",
    "javascript:alert('XSS')",
    "<iframe src=javascript:alert('XSS')>",
    "';alert('XSS');//",
    "\"><script>alert('XSS')</script>",
    "<body onload=alert('XSS')>",
]

# Advanced SQL Injection Payloads
SQL_INJECTION_PAYLOADS = [
    "' OR '1'='1",
    "' OR 1=1--",
    "admin'--",
    "' UNION SELECT NULL--",
    "1' AND SLEEP(5)--",
    "' OR '1'='1' /*",
    "'; DROP TABLE users--",
    "' OR 'x'='x",
]

# NoSQL Injection Payloads
NOSQL_PAYLOADS = [
    "{'$gt': ''}",
    "{'$ne': null}",
    "admin' || '1'=='1",
]

# Command Injection Payloads
COMMAND_INJECTION_PAYLOADS = [
    "; ls -la",
    "| whoami",
    "`id`",
    "$(whoami)",
    "&& cat /etc/passwd",
]

# LDAP Injection Payloads
LDAP_INJECTION_PAYLOADS = [
    "*",
    "*)(&",
    "*))%00",
    "admin*",
]

# SSRF Payloads
SSRF_PAYLOADS = [
    "http://127.0.0.1",
    "http://localhost",
    "http://169.254.169.254/latest/meta-data/",
    "http://metadata.google.internal/computeMetadata/v1/",
]

# Security Headers to Check
SECURITY_HEADERS = {
    "Strict-Transport-Security": "HSTS",
    "X-Frame-Options": "Clickjacking Protection",
    "X-Content-Type-Options": "MIME Sniffing Protection",
    "Content-Security-Policy": "CSP",
    "X-XSS-Protection": "XSS Filter",
    "Referrer-Policy": "Referrer Control",
    "Permissions-Policy": "Feature Policy",
    "X-Permitted-Cross-Domain-Policies": "Cross-Domain Policy",
}


# ============================================================================
# CVSS SCORING SYSTEM
# ============================================================================

class CVSSScorer:
    """CVSS v3.1 Scoring System"""

    SEVERITY_RATINGS = {
        (0.0, 0.0): "None",
        (0.1, 3.9): "Low",
        (4.0, 6.9): "Medium",
        (7.0, 8.9): "High",
        (9.0, 10.0): "Critical"
    }

    VULNERABILITY_SCORES = {
        "XSS": 6.1,
        "SQL Injection": 9.8,
        "SSRF": 8.6,
        "Command Injection": 9.8,
        "Authentication Failure": 7.5,
        "Broken Access Control": 7.5,
        "Security Misconfiguration": 5.3,
        "Vulnerable Components": 9.0,
        "Missing Security Headers": 4.0,
        "Weak SSL/TLS": 5.9,
        "CSRF": 6.5,
        "Insecure Deserialization": 8.1,
    }

    @staticmethod
    def get_severity(score: float) -> str:
        """Get severity rating from CVSS score"""
        for (low, high), severity in CVSSScorer.SEVERITY_RATINGS.items():
            if low <= score <= high:
                return severity
        return "Unknown"

    @staticmethod
    def get_score(vulnerability_type: str) -> float:
        """Get CVSS score for vulnerability type"""
        return CVSSScorer.VULNERABILITY_SCORES.get(vulnerability_type, 5.0)


# ============================================================================
# RESULT DATA STRUCTURES
# ============================================================================

class ScanResult:
    """Container for scan results with metadata"""

    def __init__(self, test_name: str, status: str, message: str,
                 severity: str = "Info", cvss_score: float = 0.0,
                 details: Optional[Dict] = None):
        self.test_name = test_name
        self.status = status  # "PASS", "FAIL", "ERROR", "SKIPPED"
        self.message = message
        self.severity = severity
        self.cvss_score = cvss_score
        self.details = details or {}
        self.timestamp = datetime.now().isoformat()

    def to_dict(self) -> Dict:
        """Convert to dictionary for JSON export"""
        return {
            "test_name": self.test_name,
            "status": self.status,
            "message": self.message,
            "severity": self.severity,
            "cvss_score": self.cvss_score,
            "details": self.details,
            "timestamp": self.timestamp
        }

    def __str__(self) -> str:
        color = {
            "PASS": Fore.GREEN,
            "FAIL": Fore.RED,
            "ERROR": Fore.YELLOW,
            "SKIPPED": Fore.CYAN
        }.get(self.status, "")

        return f"{color}[{self.test_name}] {self.status}: {self.message}{Style.RESET_ALL}"


# ============================================================================
# CORE SECURITY TESTING ENGINE
# ============================================================================

class SecurityScanner:
    """Advanced Security Scanner with OWASP Top 10 Coverage"""

    def __init__(self, target_url: str, headers: Optional[Dict] = None,
                 timeout: int = 10, max_workers: int = 5):
        self.target_url = target_url
        self.headers = headers or {"User-Agent": USER_AGENT}
        self.timeout = timeout
        self.max_workers = max_workers
        self.results: List[ScanResult] = []
        self.session = requests.Session()
        self.session.headers.update(self.headers)

    # ========================================================================
    # OWASP A01:2021 - Broken Access Control
    # ========================================================================

    def check_access_control(self) -> ScanResult:
        """Test for broken access control"""
        try:
            response = self.session.get(self.target_url, timeout=self.timeout)

            if response.status_code == 200:
                return ScanResult(
                    "Broken Access Control",
                    "PASS",
                    f"URL is publicly accessible (Status: {response.status_code})",
                    "Info",
                    0.0,
                    {"status_code": response.status_code, "url": self.target_url}
                )
            else:
                return ScanResult(
                    "Broken Access Control",
                    "PASS",
                    f"URL returned status {response.status_code}",
                    "Info",
                    0.0
                )
        except requests.RequestException as e:
            return ScanResult(
                "Broken Access Control",
                "ERROR",
                f"Error accessing {self.target_url}: {str(e)}",
                "Info"
            )

    # ========================================================================
    # OWASP A02:2021 - Cryptographic Failures
    # ========================================================================

    def check_ssl_configuration(self) -> ScanResult:
        """Test SSL/TLS configuration and certificate"""
        parsed_url = urlparse(self.target_url)
        hostname = parsed_url.hostname

        if not hostname:
            return ScanResult(
                "Cryptographic Failures",
                "ERROR",
                "Invalid hostname",
                "Info"
            )

        port = 443
        try:
            context = ssl.create_default_context()
            with socket.create_connection((hostname, port), timeout=self.timeout) as sock:
                with context.wrap_socket(sock, server_hostname=hostname) as ssock:
                    cert = ssock.getpeercert()
                    cipher = ssock.cipher()
                    version = ssock.version()

                    details = {
                        "cipher": cipher,
                        "protocol": version,
                        "cert_subject": dict(x[0] for x in cert.get('subject', [])),
                        "cert_issuer": dict(x[0] for x in cert.get('issuer', []))
                    }

                    # Check for weak protocols
                    if version in ["TLSv1", "TLSv1.1", "SSLv2", "SSLv3"]:
                        return ScanResult(
                            "Cryptographic Failures",
                            "FAIL",
                            f"Weak TLS version detected: {version}",
                            "High",
                            CVSSScorer.get_score("Weak SSL/TLS"),
                            details
                        )

                    return ScanResult(
                        "Cryptographic Failures",
                        "PASS",
                        f"SSL certificate valid for {hostname} (Protocol: {version})",
                        "Info",
                        0.0,
                        details
                    )
        except ssl.SSLError as e:
            return ScanResult(
                "Cryptographic Failures",
                "FAIL",
                f"SSL Error for {hostname}: {str(e)}",
                "High",
                CVSSScorer.get_score("Weak SSL/TLS")
            )
        except Exception as e:
            return ScanResult(
                "Cryptographic Failures",
                "ERROR",
                f"Error connecting to {hostname}: {str(e)}",
                "Info"
            )

    # ========================================================================
    # OWASP A03:2021 - Injection
    # ========================================================================

    def test_sql_injection(self) -> List[ScanResult]:
        """Test for SQL injection vulnerabilities"""
        results = []

        for payload in SQL_INJECTION_PAYLOADS:
            try:
                response = self.session.get(
                    self.target_url,
                    params={"input": payload, "id": payload, "search": payload},
                    timeout=self.timeout
                )

                # Check for SQL error patterns
                sql_errors = [
                    "sql syntax",
                    "mysql_fetch",
                    "ora-01",
                    "postgresql",
                    "sqlite_",
                    "odbc_",
                    "jdbc",
                    "warning: mysql"
                ]

                response_lower = response.text.lower()
                if any(error in response_lower for error in sql_errors):
                    results.append(ScanResult(
                        "SQL Injection",
                        "FAIL",
                        f"Potential SQL injection detected with payload: {payload}",
                        "Critical",
                        CVSSScorer.get_score("SQL Injection"),
                        {"payload": payload, "evidence": response.text[:200]}
                    ))
                    break
            except requests.RequestException:
                continue

        if not results:
            results.append(ScanResult(
                "SQL Injection",
                "PASS",
                "No SQL injection vulnerabilities detected",
                "Info"
            ))

        return results

    def test_nosql_injection(self) -> ScanResult:
        """Test for NoSQL injection"""
        for payload in NOSQL_PAYLOADS:
            try:
                response = self.session.post(
                    self.target_url,
                    json={"username": payload, "password": payload},
                    timeout=self.timeout
                )

                if response.status_code == 200 and "authenticated" in response.text.lower():
                    return ScanResult(
                        "NoSQL Injection",
                        "FAIL",
                        f"Potential NoSQL injection detected with payload: {payload}",
                        "Critical",
                        CVSSScorer.get_score("SQL Injection"),
                        {"payload": payload}
                    )
            except requests.RequestException:
                continue

        return ScanResult(
            "NoSQL Injection",
            "PASS",
            "No NoSQL injection vulnerabilities detected",
            "Info"
        )

    def test_command_injection(self) -> ScanResult:
        """Test for command injection"""
        for payload in COMMAND_INJECTION_PAYLOADS:
            try:
                response = self.session.get(
                    self.target_url,
                    params={"cmd": payload, "exec": payload},
                    timeout=self.timeout
                )

                # Check for command execution indicators
                indicators = ["root:", "uid=", "gid=", "groups=", "/bin/", "/etc/passwd"]
                if any(indicator in response.text for indicator in indicators):
                    return ScanResult(
                        "Command Injection",
                        "FAIL",
                        f"Potential command injection detected with payload: {payload}",
                        "Critical",
                        CVSSScorer.get_score("Command Injection"),
                        {"payload": payload}
                    )
            except requests.RequestException:
                continue

        return ScanResult(
            "Command Injection",
            "PASS",
            "No command injection vulnerabilities detected",
            "Info"
        )

    # ========================================================================
    # OWASP A03:2021 - XSS (Cross-Site Scripting)
    # ========================================================================

    def test_xss(self) -> List[ScanResult]:
        """Test for XSS vulnerabilities (Reflected, Stored, DOM-based)"""
        results = []

        for payload in XSS_PAYLOADS:
            try:
                # Test reflected XSS
                response = self.session.get(
                    self.target_url,
                    params={"q": payload, "search": payload, "input": payload},
                    timeout=self.timeout
                )

                if payload in response.text:
                    results.append(ScanResult(
                        "Cross-Site Scripting (XSS)",
                        "FAIL",
                        f"Potential reflected XSS detected with payload: {payload}",
                        "High",
                        CVSSScorer.get_score("XSS"),
                        {"payload": payload, "type": "reflected"}
                    ))
                    break

                # Check for DOM-based XSS indicators
                if BeautifulSoup:
                    soup = BeautifulSoup(response.text, 'html.parser')
                    scripts = soup.find_all('script')
                    for script in scripts:
                        if script.string and any(dangerous in script.string for dangerous in
                                                ['document.write', 'innerHTML', 'eval(']):
                            results.append(ScanResult(
                                "Cross-Site Scripting (XSS)",
                                "FAIL",
                                "Potential DOM-based XSS detected (dangerous JavaScript patterns)",
                                "High",
                                CVSSScorer.get_score("XSS"),
                                {"type": "dom-based"}
                            ))
                            break
            except requests.RequestException:
                continue

        if not results:
            results.append(ScanResult(
                "Cross-Site Scripting (XSS)",
                "PASS",
                "No XSS vulnerabilities detected",
                "Info"
            ))

        return results

    # ========================================================================
    # OWASP A04:2021 - Insecure Design (not directly testable, informational)
    # ========================================================================

    # ========================================================================
    # OWASP A05:2021 - Security Misconfiguration
    # ========================================================================

    def check_security_headers(self) -> List[ScanResult]:
        """Check for missing security headers"""
        results = []

        try:
            response = self.session.get(self.target_url, timeout=self.timeout)
            headers = response.headers

            missing_headers = []
            present_headers = []

            for header, description in SECURITY_HEADERS.items():
                if header not in headers:
                    missing_headers.append(f"{header} ({description})")
                else:
                    present_headers.append(f"{header}: {headers[header]}")

            if missing_headers:
                results.append(ScanResult(
                    "Security Headers",
                    "FAIL",
                    f"Missing security headers: {', '.join(missing_headers)}",
                    "Low",
                    CVSSScorer.get_score("Missing Security Headers"),
                    {"missing": missing_headers, "present": present_headers}
                ))
            else:
                results.append(ScanResult(
                    "Security Headers",
                    "PASS",
                    "All recommended security headers present",
                    "Info",
                    0.0,
                    {"present": present_headers}
                ))
        except requests.RequestException as e:
            results.append(ScanResult(
                "Security Headers",
                "ERROR",
                f"Error checking headers: {str(e)}",
                "Info"
            ))

        return results

    def check_security_misconfigurations(self) -> ScanResult:
        """Check for security misconfigurations via port scanning"""
        parsed_url = urlparse(self.target_url)
        hostname = parsed_url.hostname or "127.0.0.1"

        try:
            result = subprocess.run(
                ["nmap", "-p", "1-1000", "--open", hostname],
                capture_output=True,
                text=True,
                timeout=60
            )

            open_ports = []
            for line in result.stdout.split('\n'):
                if '/tcp' in line and 'open' in line:
                    open_ports.append(line.strip())

            if open_ports:
                return ScanResult(
                    "Security Misconfiguration (Ports)",
                    "FAIL",
                    f"Open ports detected on {hostname}",
                    "Medium",
                    CVSSScorer.get_score("Security Misconfiguration"),
                    {"open_ports": open_ports}
                )
            else:
                return ScanResult(
                    "Security Misconfiguration (Ports)",
                    "PASS",
                    f"No unusual open ports detected on {hostname}",
                    "Info"
                )
        except FileNotFoundError:
            return ScanResult(
                "Security Misconfiguration (Ports)",
                "ERROR",
                "Nmap not installed",
                "Info"
            )
        except subprocess.TimeoutExpired:
            return ScanResult(
                "Security Misconfiguration (Ports)",
                "ERROR",
                "Port scan timed out",
                "Info"
            )

    # ========================================================================
    # OWASP A06:2021 - Vulnerable and Outdated Components
    # ========================================================================

    def check_outdated_components(self) -> ScanResult:
        """Check for vulnerable Python dependencies"""
        try:
            result = subprocess.run(
                ["safety", "check", "--json"],
                capture_output=True,
                text=True,
                timeout=30
            )

            vulnerabilities = json.loads(result.stdout) if result.stdout else []

            if vulnerabilities:
                vuln_list = []
                for issue in vulnerabilities:
                    vuln_list.append(
                        f"{issue.get('package_name', 'Unknown')} "
                        f"{issue.get('installed_version', '?')}: "
                        f"{issue.get('vulnerability', 'No description')}"
                    )

                return ScanResult(
                    "Vulnerable Components",
                    "FAIL",
                    f"Found {len(vulnerabilities)} vulnerable components",
                    "High",
                    CVSSScorer.get_score("Vulnerable Components"),
                    {"vulnerabilities": vuln_list}
                )
            else:
                return ScanResult(
                    "Vulnerable Components",
                    "PASS",
                    "No vulnerable components detected",
                    "Info"
                )
        except FileNotFoundError:
            return ScanResult(
                "Vulnerable Components",
                "ERROR",
                "Safety not installed",
                "Info"
            )
        except (json.JSONDecodeError, subprocess.TimeoutExpired):
            return ScanResult(
                "Vulnerable Components",
                "ERROR",
                "Error checking components",
                "Info"
            )

    # ========================================================================
    # OWASP A07:2021 - Identification and Authentication Failures
    # ========================================================================

    def check_authentication_failures(self) -> List[ScanResult]:
        """Test for authentication and session management issues"""
        results = []

        try:
            response = self.session.get(self.target_url, timeout=self.timeout)

            # Check for session cookies
            session_cookies = [cookie for cookie in response.cookies if
                             any(name in cookie.name.lower() for name in
                             ['session', 'sess', 'token', 'auth'])]

            insecure_cookies = []
            for cookie in session_cookies:
                if not cookie.secure:
                    insecure_cookies.append(f"{cookie.name} (missing Secure flag)")
                if not cookie.has_nonstandard_attr('HttpOnly'):
                    insecure_cookies.append(f"{cookie.name} (missing HttpOnly flag)")

            if insecure_cookies:
                results.append(ScanResult(
                    "Authentication - Insecure Cookies",
                    "FAIL",
                    f"Insecure cookie configuration: {', '.join(insecure_cookies)}",
                    "Medium",
                    CVSSScorer.get_score("Authentication Failure"),
                    {"insecure_cookies": insecure_cookies}
                ))
            else:
                results.append(ScanResult(
                    "Authentication - Secure Cookies",
                    "PASS",
                    "Session cookies properly configured",
                    "Info"
                ))

            # Check for weak password policy indicators
            if BeautifulSoup and 'login' in self.target_url.lower():
                soup = BeautifulSoup(response.text, 'html.parser')
                password_fields = soup.find_all('input', {'type': 'password'})

                if password_fields:
                    has_requirements = bool(soup.find(text=re.compile(
                        r'(password.*requirements|minimum.*characters|uppercase|lowercase|special.*character)',
                        re.IGNORECASE
                    )))

                    if not has_requirements:
                        results.append(ScanResult(
                            "Authentication - Password Policy",
                            "FAIL",
                            "No visible password complexity requirements",
                            "Low",
                            4.0
                        ))

        except requests.RequestException as e:
            results.append(ScanResult(
                "Authentication Failures",
                "ERROR",
                f"Error testing authentication: {str(e)}",
                "Info"
            ))

        if not results:
            results.append(ScanResult(
                "Authentication Failures",
                "PASS",
                "No obvious authentication issues detected",
                "Info"
            ))

        return results

    # ========================================================================
    # OWASP A08:2021 - Software and Data Integrity Failures
    # ========================================================================

    def check_integrity_failures(self) -> List[ScanResult]:
        """Check for software and data integrity issues"""
        results = []

        try:
            response = self.session.get(self.target_url, timeout=self.timeout)

            if BeautifulSoup:
                soup = BeautifulSoup(response.text, 'html.parser')

                # Check for unsigned resources from CDNs
                scripts = soup.find_all('script', src=True)
                unsigned_scripts = []

                for script in scripts:
                    src = script.get('src', '')
                    if any(cdn in src for cdn in ['cdn.', 'ajax.', 'cloudflare']):
                        if not script.get('integrity'):
                            unsigned_scripts.append(src)

                if unsigned_scripts:
                    results.append(ScanResult(
                        "Integrity - Unsigned Resources",
                        "FAIL",
                        f"Found {len(unsigned_scripts)} unsigned CDN resources (missing SRI)",
                        "Medium",
                        CVSSScorer.get_score("Insecure Deserialization"),
                        {"unsigned_resources": unsigned_scripts[:5]}
                    ))
                else:
                    results.append(ScanResult(
                        "Integrity - Resource Integrity",
                        "PASS",
                        "CDN resources properly signed with SRI",
                        "Info"
                    ))

        except requests.RequestException as e:
            results.append(ScanResult(
                "Software/Data Integrity",
                "ERROR",
                f"Error checking integrity: {str(e)}",
                "Info"
            ))

        if not results:
            results.append(ScanResult(
                "Software/Data Integrity",
                "PASS",
                "No integrity issues detected",
                "Info"
            ))

        return results

    # ========================================================================
    # OWASP A09:2021 - Security Logging and Monitoring Failures
    # ========================================================================

    def check_logging_monitoring(self) -> List[ScanResult]:
        """Check for security logging and monitoring"""
        results = []

        try:
            # Test if application logs security events
            # Attempt various attacks and check for different responses
            normal_response = self.session.get(self.target_url, timeout=self.timeout)
            attack_response = self.session.get(
                self.target_url,
                params={"id": "' OR 1=1--"},
                timeout=self.timeout
            )

            # Check if error messages expose sensitive info
            error_indicators = [
                "stack trace",
                "exception",
                "debug",
                "warning:",
                "error:",
                "line ",
                "file ",
                "traceback"
            ]

            verbose_errors = any(indicator in attack_response.text.lower()
                               for indicator in error_indicators)

            if verbose_errors:
                results.append(ScanResult(
                    "Logging - Verbose Errors",
                    "FAIL",
                    "Application exposes verbose error messages",
                    "Low",
                    3.0,
                    {"evidence": attack_response.text[:200]}
                ))
            else:
                results.append(ScanResult(
                    "Logging - Error Handling",
                    "PASS",
                    "Error messages properly sanitized",
                    "Info"
                ))

        except requests.RequestException as e:
            results.append(ScanResult(
                "Logging and Monitoring",
                "ERROR",
                f"Error testing logging: {str(e)}",
                "Info"
            ))

        if not results:
            results.append(ScanResult(
                "Logging and Monitoring",
                "PASS",
                "No logging issues detected",
                "Info"
            ))

        return results

    # ========================================================================
    # OWASP A10:2021 - Server-Side Request Forgery (SSRF)
    # ========================================================================

    def test_ssrf(self) -> List[ScanResult]:
        """Test for SSRF vulnerabilities"""
        results = []

        for payload in SSRF_PAYLOADS:
            try:
                response = self.session.get(
                    self.target_url,
                    params={"url": payload, "fetch": payload, "proxy": payload},
                    timeout=self.timeout
                )

                # Check for SSRF indicators
                ssrf_indicators = [
                    "169.254.169.254",
                    "metadata",
                    "localhost",
                    "127.0.0.1"
                ]

                if any(indicator in response.text for indicator in ssrf_indicators):
                    results.append(ScanResult(
                        "Server-Side Request Forgery (SSRF)",
                        "FAIL",
                        f"Potential SSRF detected with payload: {payload}",
                        "High",
                        CVSSScorer.get_score("SSRF"),
                        {"payload": payload}
                    ))
                    break
            except requests.RequestException:
                continue

        if not results:
            results.append(ScanResult(
                "Server-Side Request Forgery (SSRF)",
                "PASS",
                "No SSRF vulnerabilities detected",
                "Info"
            ))

        return results

    # ========================================================================
    # API SECURITY TESTING
    # ========================================================================

    def test_api_security(self) -> List[ScanResult]:
        """Test API-specific security issues"""
        results = []

        try:
            response = self.session.options(self.target_url, timeout=self.timeout)

            # Check CORS configuration
            cors_header = response.headers.get('Access-Control-Allow-Origin')
            if cors_header == '*':
                results.append(ScanResult(
                    "API - CORS Misconfiguration",
                    "FAIL",
                    "Overly permissive CORS policy (allows all origins)",
                    "Medium",
                    5.0,
                    {"cors_header": cors_header}
                ))

            # Check for GraphQL introspection
            if 'graphql' in self.target_url.lower():
                graphql_query = {"query": "{__schema{types{name}}}"}
                graphql_response = self.session.post(
                    self.target_url,
                    json=graphql_query,
                    timeout=self.timeout
                )

                if graphql_response.status_code == 200 and '__schema' in graphql_response.text:
                    results.append(ScanResult(
                        "API - GraphQL Introspection",
                        "FAIL",
                        "GraphQL introspection enabled (information disclosure)",
                        "Low",
                        3.0
                    ))

            # Check for API versioning
            headers = response.headers
            if 'api-version' not in headers.get('content-type', '').lower():
                results.append(ScanResult(
                    "API - Versioning",
                    "FAIL",
                    "No API versioning detected",
                    "Low",
                    2.0
                ))

        except requests.RequestException as e:
            results.append(ScanResult(
                "API Security",
                "ERROR",
                f"Error testing API security: {str(e)}",
                "Info"
            ))

        if not results:
            results.append(ScanResult(
                "API Security",
                "PASS",
                "No API security issues detected",
                "Info"
            ))

        return results

    # ========================================================================
    # MAIN SCAN ORCHESTRATION
    # ========================================================================

    def run_all_tests(self, exclusions: List[str] = None,
                     exclude_hackerone: bool = False,
                     use_threading: bool = True) -> List[ScanResult]:
        """Run all security tests"""
        exclusions = exclusions or []
        all_results = []

        # Define all test methods
        test_methods = [
            ("Broken Access Control", self.check_access_control),
            ("Cryptographic Failures", self.check_ssl_configuration),
            ("SQL Injection", lambda: self.test_sql_injection()),
            ("NoSQL Injection", self.test_nosql_injection),
            ("Command Injection", self.test_command_injection),
            ("Cross-Site Scripting (XSS)", lambda: self.test_xss()),
            ("Security Headers", lambda: self.check_security_headers()),
            ("Security Misconfiguration", self.check_security_misconfigurations),
            ("Vulnerable Components", self.check_outdated_components),
            ("Authentication Failures", lambda: self.check_authentication_failures()),
            ("Software/Data Integrity", lambda: self.check_integrity_failures()),
            ("Logging and Monitoring", lambda: self.check_logging_monitoring()),
            ("SSRF", lambda: self.test_ssrf()),
            ("API Security", lambda: self.test_api_security()),
        ]

        def run_test(test_name, test_func):
            """Run a single test with exclusion logic"""
            # Check exclusions
            if test_name in exclusions:
                return ScanResult(test_name, "SKIPPED", "In exclusions list", "Info")

            if exclude_hackerone and should_exclude_test(test_name, exclude_hackerone):
                return ScanResult(
                    test_name,
                    "SKIPPED",
                    "Excluded by HackerOne Core Ineligible Findings",
                    "Info"
                )

            try:
                result = test_func()
                return result if isinstance(result, ScanResult) else result
            except Exception as e:
                return ScanResult(test_name, "ERROR", f"Test error: {str(e)}", "Info")

        if use_threading and self.max_workers > 1:
            # Run tests in parallel
            with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
                futures = {
                    executor.submit(run_test, name, func): name
                    for name, func in test_methods
                }

                for future in as_completed(futures):
                    result = future.result()
                    if isinstance(result, list):
                        all_results.extend(result)
                    else:
                        all_results.append(result)
        else:
            # Run tests sequentially
            for test_name, test_func in test_methods:
                result = run_test(test_name, test_func)
                if isinstance(result, list):
                    all_results.extend(result)
                else:
                    all_results.append(result)

        self.results = all_results
        return all_results


# ============================================================================
# REPORT GENERATION
# ============================================================================

class ReportGenerator:
    """Generate reports in multiple formats"""

    @staticmethod
    def generate_json(results: List[ScanResult], output_file: str, metadata: Dict = None):
        """Generate JSON report"""
        report = {
            "metadata": metadata or {},
            "scan_timestamp": datetime.now().isoformat(),
            "total_tests": len(results),
            "summary": ReportGenerator._generate_summary(results),
            "results": [r.to_dict() for r in results]
        }

        with open(output_file, 'w') as f:
            json.dump(report, f, indent=2)

        return output_file

    @staticmethod
    def generate_pdf(results: List[ScanResult], output_file: str, metadata: Dict = None):
        """Generate PDF report"""
        if not SimpleDocTemplate:
            raise ImportError("reportlab not installed")

        doc = SimpleDocTemplate(output_file, pagesize=letter)
        story = []
        styles = getSampleStyleSheet()

        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#2C3E50'),
            spaceAfter=30,
        )
        story.append(Paragraph("OWASP QuickCheck Security Report", title_style))
        story.append(Spacer(1, 12))

        # Metadata
        if metadata:
            story.append(Paragraph(f"<b>Target:</b> {metadata.get('target_url', 'N/A')}", styles['Normal']))
            story.append(Paragraph(f"<b>Scan Date:</b> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
            story.append(Spacer(1, 20))

        # Executive Summary
        summary = ReportGenerator._generate_summary(results)
        story.append(Paragraph("<b>Executive Summary</b>", styles['Heading2']))
        story.append(Paragraph(f"Total Tests: {summary['total_tests']}", styles['Normal']))
        story.append(Paragraph(f"Critical: {summary['critical']}", styles['Normal']))
        story.append(Paragraph(f"High: {summary['high']}", styles['Normal']))
        story.append(Paragraph(f"Medium: {summary['medium']}", styles['Normal']))
        story.append(Paragraph(f"Low: {summary['low']}", styles['Normal']))
        story.append(Spacer(1, 20))

        # Results table
        story.append(Paragraph("<b>Detailed Findings</b>", styles['Heading2']))
        story.append(Spacer(1, 12))

        table_data = [['Test', 'Status', 'Severity', 'CVSS', 'Message']]
        for result in results:
            table_data.append([
                Paragraph(result.test_name, styles['Normal']),
                result.status,
                result.severity,
                str(result.cvss_score),
                Paragraph(result.message[:100], styles['Normal'])
            ])

        t = Table(table_data, colWidths=[2*inch, 0.8*inch, 0.8*inch, 0.6*inch, 3*inch])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(t)

        doc.build(story)
        return output_file

    @staticmethod
    def generate_html(results: List[ScanResult], output_file: str, metadata: Dict = None):
        """Generate HTML report"""
        if not Template:
            raise ImportError("jinja2 not installed")

        html_template = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>OWASP QuickCheck Security Report</title>
    <style>
        body { font-family: Arial, sans-serif; margin: 20px; background: #f5f5f5; }
        .container { max-width: 1200px; margin: 0 auto; background: white; padding: 30px; border-radius: 8px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }
        h1 { color: #2C3E50; border-bottom: 3px solid #3498DB; padding-bottom: 10px; }
        h2 { color: #34495E; margin-top: 30px; }
        .summary { display: grid; grid-template-columns: repeat(auto-fit, minmax(150px, 1fr)); gap: 15px; margin: 20px 0; }
        .summary-card { background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 20px; border-radius: 8px; text-align: center; }
        .summary-card h3 { margin: 0; font-size: 2em; }
        .summary-card p { margin: 5px 0 0 0; }
        table { width: 100%; border-collapse: collapse; margin: 20px 0; }
        th, td { padding: 12px; text-align: left; border-bottom: 1px solid #ddd; }
        th { background-color: #3498DB; color: white; }
        tr:hover { background-color: #f5f5f5; }
        .status-PASS { color: green; font-weight: bold; }
        .status-FAIL { color: red; font-weight: bold; }
        .status-ERROR { color: orange; font-weight: bold; }
        .status-SKIPPED { color: gray; font-weight: bold; }
        .severity-Critical { background-color: #e74c3c; color: white; padding: 4px 8px; border-radius: 4px; }
        .severity-High { background-color: #e67e22; color: white; padding: 4px 8px; border-radius: 4px; }
        .severity-Medium { background-color: #f39c12; color: white; padding: 4px 8px; border-radius: 4px; }
        .severity-Low { background-color: #3498db; color: white; padding: 4px 8px; border-radius: 4px; }
        .severity-Info { background-color: #95a5a6; color: white; padding: 4px 8px; border-radius: 4px; }
        .metadata { background: #ecf0f1; padding: 15px; border-radius: 8px; margin: 20px 0; }
    </style>
</head>
<body>
    <div class="container">
        <h1>🔒 OWASP QuickCheck Security Report</h1>

        <div class="metadata">
            <strong>Target URL:</strong> {{ metadata.target_url }}<br>
            <strong>Scan Date:</strong> {{ metadata.scan_date }}<br>
            <strong>Scanner Version:</strong> 2.0
        </div>

        <h2>Executive Summary</h2>
        <div class="summary">
            <div class="summary-card">
                <h3>{{ summary.total_tests }}</h3>
                <p>Total Tests</p>
            </div>
            <div class="summary-card" style="background: linear-gradient(135deg, #e74c3c 0%, #c0392b 100%);">
                <h3>{{ summary.critical }}</h3>
                <p>Critical</p>
            </div>
            <div class="summary-card" style="background: linear-gradient(135deg, #e67e22 0%, #d35400 100%);">
                <h3>{{ summary.high }}</h3>
                <p>High</p>
            </div>
            <div class="summary-card" style="background: linear-gradient(135deg, #f39c12 0%, #f1c40f 100%);">
                <h3>{{ summary.medium }}</h3>
                <p>Medium</p>
            </div>
            <div class="summary-card" style="background: linear-gradient(135deg, #3498db 0%, #2980b9 100%);">
                <h3>{{ summary.low }}</h3>
                <p>Low</p>
            </div>
        </div>

        <h2>Detailed Findings</h2>
        <table>
            <thead>
                <tr>
                    <th>Test Name</th>
                    <th>Status</th>
                    <th>Severity</th>
                    <th>CVSS Score</th>
                    <th>Message</th>
                </tr>
            </thead>
            <tbody>
                {% for result in results %}
                <tr>
                    <td>{{ result.test_name }}</td>
                    <td class="status-{{ result.status }}">{{ result.status }}</td>
                    <td><span class="severity-{{ result.severity }}">{{ result.severity }}</span></td>
                    <td>{{ result.cvss_score }}</td>
                    <td>{{ result.message }}</td>
                </tr>
                {% endfor %}
            </tbody>
        </table>
    </div>
</body>
</html>
        """

        template = Template(html_template)
        summary = ReportGenerator._generate_summary(results)

        html_content = template.render(
            metadata={
                'target_url': metadata.get('target_url', 'N/A') if metadata else 'N/A',
                'scan_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            },
            summary=summary,
            results=results
        )

        with open(output_file, 'w') as f:
            f.write(html_content)

        return output_file

    @staticmethod
    def generate_excel(results: List[ScanResult], output_file: str, metadata: Dict = None):
        """Generate Excel report"""
        if not openpyxl:
            raise ImportError("openpyxl not installed")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Security Scan Results"

        # Headers
        headers = ['Test Name', 'Status', 'Severity', 'CVSS Score', 'Message', 'Timestamp']
        ws.append(headers)

        # Style headers
        header_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Add data
        for result in results:
            ws.append([
                result.test_name,
                result.status,
                result.severity,
                result.cvss_score,
                result.message,
                result.timestamp
            ])

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(output_file)
        return output_file

    @staticmethod
    def _generate_summary(results: List[ScanResult]) -> Dict:
        """Generate summary statistics"""
        summary = {
            "total_tests": len(results),
            "passed": sum(1 for r in results if r.status == "PASS"),
            "failed": sum(1 for r in results if r.status == "FAIL"),
            "errors": sum(1 for r in results if r.status == "ERROR"),
            "skipped": sum(1 for r in results if r.status == "SKIPPED"),
            "critical": sum(1 for r in results if r.severity == "Critical"),
            "high": sum(1 for r in results if r.severity == "High"),
            "medium": sum(1 for r in results if r.severity == "Medium"),
            "low": sum(1 for r in results if r.severity == "Low"),
        }
        return summary


# ============================================================================
# CLI INTERFACE
# ============================================================================

def cli_mode():
    """Command-line interface for CI/CD integration"""
    parser = argparse.ArgumentParser(
        description="OWASP QuickCheck - Advanced Security Testing Tool",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  %(prog)s -u https://example.com
  %(prog)s -u https://example.com --json report.json
  %(prog)s -u https://example.com --pdf report.pdf --html report.html
  %(prog)s -u https://example.com --exclude-hackerone --threads 10
  %(prog)s -u https://example.com --severity-threshold high --exit-code
        """
    )

    parser.add_argument('-u', '--url', required=True, help='Target URL to scan')
    parser.add_argument('--json', metavar='FILE', help='Output JSON report')
    parser.add_argument('--pdf', metavar='FILE', help='Output PDF report')
    parser.add_argument('--html', metavar='FILE', help='Output HTML report')
    parser.add_argument('--excel', metavar='FILE', help='Output Excel report')
    parser.add_argument('--exclude-hackerone', action='store_true',
                       help='Exclude HackerOne Core Ineligible Findings')
    parser.add_argument('--exclusions', metavar='FILE', default='exclusions.txt',
                       help='File containing test exclusions')
    parser.add_argument('--threads', type=int, default=5, metavar='N',
                       help='Number of threads for parallel scanning (default: 5)')
    parser.add_argument('--timeout', type=int, default=10, metavar='SEC',
                       help='Request timeout in seconds (default: 10)')
    parser.add_argument('--header', action='append', metavar='KEY:VALUE',
                       help='Custom HTTP header (can be used multiple times)')
    parser.add_argument('--severity-threshold', choices=['low', 'medium', 'high', 'critical'],
                       help='Minimum severity to report')
    parser.add_argument('--exit-code', action='store_true',
                       help='Exit with code 1 if vulnerabilities found')
    parser.add_argument('--verbose', '-v', action='store_true',
                       help='Verbose output')
    parser.add_argument('--no-color', action='store_true',
                       help='Disable colored output')

    args = parser.parse_args()

    # Parse custom headers
    headers = {}
    if args.header:
        for header in args.header:
            if ':' in header:
                key, value = header.split(':', 1)
                headers[key.strip()] = value.strip()

    # Load exclusions
    exclusions = []
    if os.path.exists(args.exclusions):
        with open(args.exclusions, 'r') as f:
            exclusions = [line.strip() for line in f if line.strip() and not line.startswith('#')]

    # Print banner
    print(f"{Fore.CYAN}{Style.BRIGHT}")
    print("=" * 70)
    print("   OWASP QuickCheck v2.0 - Advanced Security Scanner")
    print("=" * 70)
    print(f"{Style.RESET_ALL}")
    print(f"Target: {args.url}")
    print(f"Threads: {args.threads}")
    print(f"Timeout: {args.timeout}s")
    if args.exclude_hackerone:
        print(f"{Fore.YELLOW}HackerOne exclusions: Enabled{Style.RESET_ALL}")
    print("-" * 70)

    # Run scan
    scanner = SecurityScanner(args.url, headers, args.timeout, args.threads)

    start_time = time.time()
    results = scanner.run_all_tests(exclusions, args.exclude_hackerone, use_threading=(args.threads > 1))
    elapsed_time = time.time() - start_time

    # Filter by severity threshold
    if args.severity_threshold:
        severity_order = {'low': 1, 'medium': 2, 'high': 3, 'critical': 4}
        threshold = severity_order[args.severity_threshold]
        results = [r for r in results if severity_order.get(r.severity.lower(), 0) >= threshold]

    # Print results
    print(f"\n{Fore.CYAN}{Style.BRIGHT}Scan Results:{Style.RESET_ALL}\n")
    for result in results:
        print(result)

    # Generate summary
    summary = ReportGenerator._generate_summary(results)
    print(f"\n{Fore.CYAN}{Style.BRIGHT}Summary:{Style.RESET_ALL}")
    print(f"Total Tests: {summary['total_tests']}")
    print(f"{Fore.GREEN}Passed: {summary['passed']}{Style.RESET_ALL}")
    print(f"{Fore.RED}Failed: {summary['failed']}{Style.RESET_ALL}")
    print(f"{Fore.YELLOW}Errors: {summary['errors']}{Style.RESET_ALL}")
    print(f"{Fore.CYAN}Skipped: {summary['skipped']}{Style.RESET_ALL}")
    print(f"\nSeverity Breakdown:")
    print(f"{Fore.RED}Critical: {summary['critical']}{Style.RESET_ALL}")
    print(f"{Fore.YELLOW}High: {summary['high']}{Style.RESET_ALL}")
    print(f"{Fore.BLUE}Medium: {summary['medium']}{Style.RESET_ALL}")
    print(f"{Fore.CYAN}Low: {summary['low']}{Style.RESET_ALL}")
    print(f"\nElapsed Time: {elapsed_time:.2f}s")

    # Generate reports
    metadata = {'target_url': args.url}

    if args.json:
        ReportGenerator.generate_json(results, args.json, metadata)
        print(f"\n{Fore.GREEN}JSON report saved to: {args.json}{Style.RESET_ALL}")

    if args.pdf:
        try:
            ReportGenerator.generate_pdf(results, args.pdf, metadata)
            print(f"{Fore.GREEN}PDF report saved to: {args.pdf}{Style.RESET_ALL}")
        except ImportError as e:
            print(f"{Fore.YELLOW}Warning: {e}{Style.RESET_ALL}")

    if args.html:
        try:
            ReportGenerator.generate_html(results, args.html, metadata)
            print(f"{Fore.GREEN}HTML report saved to: {args.html}{Style.RESET_ALL}")
        except ImportError as e:
            print(f"{Fore.YELLOW}Warning: {e}{Style.RESET_ALL}")

    if args.excel:
        try:
            ReportGenerator.generate_excel(results, args.excel, metadata)
            print(f"{Fore.GREEN}Excel report saved to: {args.excel}{Style.RESET_ALL}")
        except ImportError as e:
            print(f"{Fore.YELLOW}Warning: {e}{Style.RESET_ALL}")

    # Exit code based on findings
    if args.exit_code and summary['failed'] > 0:
        sys.exit(1)

    sys.exit(0)


# ============================================================================
# GUI INTERFACE
# ============================================================================

class QuickCheckApp:
    """GUI Application for OWASP QuickCheck"""

    def __init__(self):
        self.root = tk.Tk()
        self.root.title("OWASP QuickCheck v2.0 - Advanced Security Scanner")
        self.root.geometry("800x600")

        # Create main frame
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        # Title
        title_label = ttk.Label(main_frame, text="OWASP QuickCheck v2.0",
                               font=('Arial', 16, 'bold'))
        title_label.grid(row=0, column=0, columnspan=2, pady=10)

        # Document selection
        ttk.Label(main_frame, text="Documentation:").grid(row=1, column=0, sticky=tk.W, pady=5)
        self.documents = self.list_documents()
        self.selected_file = tk.StringVar()
        self.dropdown = ttk.Combobox(main_frame, textvariable=self.selected_file,
                                     values=self.documents, state="readonly", width=50)
        self.dropdown.grid(row=1, column=1, pady=5)

        self.open_button = ttk.Button(main_frame, text="Open Document",
                                      command=self.open_document)
        self.open_button.grid(row=2, column=1, pady=5)

        # Options frame
        options_frame = ttk.LabelFrame(main_frame, text="Scan Options", padding="10")
        options_frame.grid(row=3, column=0, columnspan=2, pady=10, sticky=(tk.W, tk.E))

        # HackerOne exclusions
        self.exclude_hackerone = tk.BooleanVar(value=False)
        ttk.Checkbutton(options_frame, text="Exclude HackerOne Core Ineligible Findings",
                       variable=self.exclude_hackerone).grid(row=0, column=0, sticky=tk.W)

        # Threading option
        self.use_threading = tk.BooleanVar(value=True)
        ttk.Checkbutton(options_frame, text="Use Multi-threaded Scanning",
                       variable=self.use_threading).grid(row=1, column=0, sticky=tk.W)

        # Thread count
        ttk.Label(options_frame, text="Threads:").grid(row=2, column=0, sticky=tk.W, pady=5)
        self.thread_count = tk.IntVar(value=5)
        ttk.Spinbox(options_frame, from_=1, to=20, textvariable=self.thread_count,
                   width=10).grid(row=2, column=1, sticky=tk.W)

        # Buttons frame
        buttons_frame = ttk.Frame(main_frame)
        buttons_frame.grid(row=4, column=0, columnspan=2, pady=10)

        ttk.Button(buttons_frame, text="View HackerOne Exclusions",
                  command=self.view_hackerone_exclusions).grid(row=0, column=0, padx=5)

        ttk.Button(buttons_frame, text="Run Security Tests",
                  command=self.run_tests, style='Accent.TButton').grid(row=0, column=1, padx=5)

        ttk.Button(buttons_frame, text="Export Reports",
                  command=self.export_reports).grid(row=0, column=2, padx=5)

        self.root.mainloop()

    def list_documents(self):
        """Lists available HTML files in the documents folder"""
        doc_path = os.path.join(os.getcwd(), "documents")
        if not os.path.exists(doc_path):
            return []
        return [f for f in os.listdir(doc_path) if f.endswith(".html")]

    def open_document(self):
        """Opens the selected document in the default web browser"""
        selected_file = self.selected_file.get()
        if selected_file:
            file_path = os.path.join(os.getcwd(), "documents", selected_file)
            webbrowser.open(file_path)

    def view_hackerone_exclusions(self):
        """Display the HackerOne Core Ineligible Findings list"""
        exclusion_window = tk.Toplevel(self.root)
        exclusion_window.title("HackerOne Core Ineligible Findings")
        exclusion_window.geometry("600x400")

        frame = ttk.Frame(exclusion_window)
        frame.pack(fill="both", expand=True, padx=10, pady=10)

        scrollbar = ttk.Scrollbar(frame)
        scrollbar.pack(side="right", fill="y")

        text_area = tk.Text(frame, wrap="word", yscrollcommand=scrollbar.set)
        text_area.pack(fill="both", expand=True)

        scrollbar.config(command=text_area.yview)

        text_area.insert("1.0", "HackerOne Core Ineligible Findings:\n\n")
        for i, item in enumerate(get_all_exclusions(), 1):
            text_area.insert("end", f"{i}. {item}\n")

        text_area.insert("end", "\n\nExcluded Tests:\n")
        for test in get_excluded_tests():
            text_area.insert("end", f"- {test}\n")

        text_area.config(state="disabled")

    def run_tests(self):
        """Run security tests"""
        target_url = simpledialog.askstring("Target URL", "Enter the website URL to test:")
        if not target_url:
            messagebox.showinfo("OWASP QuickCheck", "No URL entered. Cancelled.")
            return

        custom_header_value = simpledialog.askstring(
            "Custom Header",
            "Enter the value for the 'Security-Research' header (optional):"
        )
        headers = {"Security-Research": custom_header_value} if custom_header_value else {}

        # Load exclusions
        exclusions = []
        if os.path.exists("exclusions.txt"):
            with open("exclusions.txt", "r") as f:
                exclusions = [line.strip() for line in f if line.strip() and not line.startswith('#')]

        # Show progress
        messagebox.showinfo("OWASP QuickCheck",
                          "Starting comprehensive security scan. This may take several minutes.")

        # Run scan
        scanner = SecurityScanner(
            target_url,
            headers,
            timeout=10,
            max_workers=self.thread_count.get()
        )

        results = scanner.run_all_tests(
            exclusions,
            self.exclude_hackerone.get(),
            self.use_threading.get()
        )

        # Store results for export
        self.last_results = results
        self.last_target_url = target_url

        # Display results
        self.show_results(results)

    def show_results(self, results: List[ScanResult]):
        """Display scan results in a new window"""
        results_window = tk.Toplevel(self.root)
        results_window.title("Scan Results")
        results_window.geometry("900x600")

        # Create treeview
        frame = ttk.Frame(results_window)
        frame.pack(fill="both", expand=True, padx=10, pady=10)

        scrollbar_y = ttk.Scrollbar(frame, orient="vertical")
        scrollbar_y.pack(side="right", fill="y")

        scrollbar_x = ttk.Scrollbar(frame, orient="horizontal")
        scrollbar_x.pack(side="bottom", fill="x")

        tree = ttk.Treeview(
            frame,
            columns=("Test", "Status", "Severity", "CVSS", "Message"),
            show="headings",
            yscrollcommand=scrollbar_y.set,
            xscrollcommand=scrollbar_x.set
        )

        tree.heading("Test", text="Test Name")
        tree.heading("Status", text="Status")
        tree.heading("Severity", text="Severity")
        tree.heading("CVSS", text="CVSS Score")
        tree.heading("Message", text="Message")

        tree.column("Test", width=200)
        tree.column("Status", width=80)
        tree.column("Severity", width=80)
        tree.column("CVSS", width=80)
        tree.column("Message", width=400)

        for result in results:
            tree.insert("", "end", values=(
                result.test_name,
                result.status,
                result.severity,
                result.cvss_score,
                result.message
            ))

        tree.pack(fill="both", expand=True)
        scrollbar_y.config(command=tree.yview)
        scrollbar_x.config(command=tree.xview)

        # Summary
        summary = ReportGenerator._generate_summary(results)
        summary_text = f"\nTotal: {summary['total_tests']} | " \
                      f"Passed: {summary['passed']} | " \
                      f"Failed: {summary['failed']} | " \
                      f"Critical: {summary['critical']} | " \
                      f"High: {summary['high']} | " \
                      f"Medium: {summary['medium']} | " \
                      f"Low: {summary['low']}"

        summary_label = ttk.Label(results_window, text=summary_text)
        summary_label.pack(pady=5)

    def export_reports(self):
        """Export last scan results to various formats"""
        if not hasattr(self, 'last_results'):
            messagebox.showerror("Error", "No scan results available. Please run a scan first.")
            return

        export_window = tk.Toplevel(self.root)
        export_window.title("Export Reports")
        export_window.geometry("400x250")

        frame = ttk.Frame(export_window, padding="20")
        frame.pack(fill="both", expand=True)

        ttk.Label(frame, text="Select export formats:", font=('Arial', 12, 'bold')).pack(pady=10)

        export_json = tk.BooleanVar(value=True)
        export_html = tk.BooleanVar(value=True)
        export_pdf = tk.BooleanVar(value=False)
        export_excel = tk.BooleanVar(value=False)

        ttk.Checkbutton(frame, text="JSON Report", variable=export_json).pack(anchor=tk.W, pady=5)
        ttk.Checkbutton(frame, text="HTML Report", variable=export_html).pack(anchor=tk.W, pady=5)
        ttk.Checkbutton(frame, text="PDF Report", variable=export_pdf).pack(anchor=tk.W, pady=5)
        ttk.Checkbutton(frame, text="Excel Report", variable=export_excel).pack(anchor=tk.W, pady=5)

        def do_export():
            metadata = {'target_url': self.last_target_url}
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

            try:
                if export_json.get():
                    ReportGenerator.generate_json(self.last_results, f"owasp_report_{timestamp}.json", metadata)
                if export_html.get():
                    ReportGenerator.generate_html(self.last_results, f"owasp_report_{timestamp}.html", metadata)
                if export_pdf.get():
                    ReportGenerator.generate_pdf(self.last_results, f"owasp_report_{timestamp}.pdf", metadata)
                if export_excel.get():
                    ReportGenerator.generate_excel(self.last_results, f"owasp_report_{timestamp}.xlsx", metadata)

                messagebox.showinfo("Success", f"Reports exported successfully with timestamp: {timestamp}")
                export_window.destroy()
            except Exception as e:
                messagebox.showerror("Export Error", f"Error exporting reports: {str(e)}")

        ttk.Button(frame, text="Export", command=do_export).pack(pady=20)


# ============================================================================
# MAIN ENTRY POINT
# ============================================================================

if __name__ == "__main__":
    if len(sys.argv) > 1:
        # CLI mode
        cli_mode()
    else:
        # GUI mode
        QuickCheckApp()
